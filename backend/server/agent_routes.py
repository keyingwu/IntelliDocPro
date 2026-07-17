"""Document Agent endpoints: CRUD, agent-bound bulk runs, stored
results and Excel export. The stateless extraction endpoints live in app.py."""

import io
import threading
import uuid

from fastapi import APIRouter, File, Form, UploadFile
from fastapi.responses import JSONResponse, Response, StreamingResponse
from pydantic import BaseModel, ConfigDict, Field

import intellidocpro
from intellidocpro.engines import ENGINES
from intellidocpro.errors import UnknownEngine

from . import jobs, store

router = APIRouter()


class AgentIn(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    name: str
    description: str = ""
    engine: str = "openai"
    model: str | None = None
    extraction_schema: intellidocpro.ExtractionSchema = Field(alias="schema")


class AgentPatch(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    name: str | None = None
    description: str | None = None
    engine: str | None = None
    model: str | None = None
    extraction_schema: intellidocpro.ExtractionSchema | None = Field(default=None, alias="schema")


def _check_engine_name(engine: str) -> None:
    if engine not in ENGINES:
        raise UnknownEngine(
            f"unknown engine '{engine}'; available: {', '.join(sorted(ENGINES))}"
        )


def _not_found(agent_id: str) -> JSONResponse:
    return JSONResponse(
        status_code=404, content={"error": "AgentNotFound", "detail": agent_id}
    )


@router.post("/agents", status_code=201)
def create_agent(body: AgentIn):
    _check_engine_name(body.engine)
    return store.create_agent(
        name=body.name,
        description=body.description,
        engine=body.engine,
        model=body.model,
        schema=body.extraction_schema.model_dump(),
    )


@router.get("/agents")
def list_agents():
    return store.list_agents()


@router.get("/agents/{agent_id}")
def get_agent(agent_id: str):
    agent = store.get_agent(agent_id)
    return agent if agent else _not_found(agent_id)


@router.put("/agents/{agent_id}")
def update_agent(agent_id: str, body: AgentPatch):
    if body.engine is not None:
        _check_engine_name(body.engine)
    fields = {
        "name": body.name,
        "description": body.description,
        "engine": body.engine,
        "model": body.model,
        "schema": body.extraction_schema.model_dump() if body.extraction_schema else None,
    }
    agent = store.update_agent(agent_id, **fields)
    return agent if agent else _not_found(agent_id)


@router.delete("/agents/{agent_id}", status_code=204)
def delete_agent(agent_id: str):
    if not store.delete_agent(agent_id):
        return _not_found(agent_id)


def _safe_filename(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_. " else "_" for c in name)


@router.get("/documents/{document_id}")
def get_document(document_id: str):
    """Serve a stored upload. The single download point for original files;
    tenant authorization lands here when multi-tenancy arrives."""
    found = store.get_document(document_id)
    if found is None:
        return JSONResponse(
            status_code=404,
            content={"error": "DocumentNotFound", "detail": document_id},
        )
    meta, data = found
    return Response(
        content=data,
        media_type=meta["content_type"] or "application/octet-stream",
        headers={
            "Content-Disposition": f'inline; filename="{_safe_filename(meta["filename"])}"'
        },
    )


@router.put("/agents/{agent_id}/sample")
async def set_sample(agent_id: str, file: UploadFile = File(...)):
    """Attach (or replace) the agent's sample document."""
    if store.get_agent(agent_id) is None:
        return _not_found(agent_id)
    data = await file.read()
    doc = store.save_document(file.filename or "document", data, file.content_type)
    return store.set_sample_document(agent_id, doc["id"])


@router.post("/agents/{agent_id}/bulk", status_code=202)
async def agent_bulk(agent_id: str, files: list[UploadFile] = File(...)):
    """Run a bulk extraction with the agent's stored schema/engine/model.
    Results are persisted when the job completes; live progress via GET /bulk/{job_id}."""
    agent = store.get_agent(agent_id)
    if agent is None:
        return _not_found(agent_id)
    intellidocpro.get_engine(agent["engine"], model=agent["model"])  # config errors now

    documents, document_ids = [], []
    for f in files:
        name = f.filename or "document"
        data = await f.read()
        documents.append((name, data))
        document_ids.append(store.save_document(name, data, f.content_type)["id"])
    job_id = uuid.uuid4().hex[:12]
    jobs.publish(
        job_id,
        intellidocpro.BulkReport(
            engine=agent["engine"],
            model=agent["model"],
            total=len(documents),
            entries=[intellidocpro.BulkFileEntry(filename=n) for n, _ in documents],
        ),
    )

    def run() -> None:
        report = intellidocpro.bulk_extract(
            documents,
            agent["schema"],
            engine=agent["engine"],
            model=agent["model"],
            on_update=lambda snap: jobs.publish(job_id, snap),
        )
        store.save_run(agent_id, report, document_ids=document_ids)

    threading.Thread(target=run, daemon=True).start()
    return {"job_id": job_id, "total": len(documents)}


@router.get("/agents/{agent_id}/results")
def list_results(agent_id: str, filter: str = "all"):
    if store.get_agent(agent_id) is None:
        return _not_found(agent_id)
    return store.list_results(agent_id, filter=filter)


@router.get("/agents/{agent_id}/export.xlsx")
def export_xlsx(agent_id: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    agent = store.get_agent(agent_id)
    if agent is None:
        return _not_found(agent_id)
    rows = store.list_results(agent_id, filter="all")
    fields = agent["schema"]["fields"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    header = ["Filename", "Status", "Needs review", *[f["name"] for f in fields],
              "Cost USD", "Extracted at"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    review_fill = PatternFill(start_color="FCF6E7", end_color="FCF6E7", fill_type="solid")
    for row in rows:
        by_field = {v["field"]: v for v in row["values"]}
        # results predating field keys stored the display name in `field`
        cell_for = [by_field.get(f.get("key") or "") or by_field.get(f["name"], {}) for f in fields]
        cells = [
            row["filename"],
            row["status"],
            "yes" if row["needs_review"] else "no",
            *[v.get("value") for v in cell_for],
            row["cost_usd"],
            row["created_at"],
        ]
        ws.append(cells)
        for offset, v in enumerate(cell_for):
            if v.get("needs_review"):
                ws.cell(row=ws.max_row, column=4 + offset).fill = review_fill

    for column_cells in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 48)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in agent["name"])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}-results.xlsx"'},
    )
