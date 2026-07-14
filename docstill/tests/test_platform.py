import io
import json
import time

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import docstill
import server.app as app_module
from docstill.bulk import BulkFileEntry, BulkReport
from docstill.pricing import CostBreakdown
from docstill.result import ExtractionResult, FieldValue

PDF = b"%PDF-1.4 fake"
SCHEMA = {"fields": [{"name": "Lieferant"}, {"name": "Betrag", "type": "amount"}]}


@pytest.fixture
def client():
    return TestClient(app_module.app)


def _create(client, **overrides):
    body = {"name": "Invoices", "engine": "claude", "schema": SCHEMA}
    body.update(overrides)
    resp = client.post("/assistants", json=body)
    assert resp.status_code == 201, resp.text
    return resp.json()


def test_assistant_crud(client):
    a = _create(client)
    assert a["name"] == "Invoices"
    assert a["doc_count"] == 0
    assert [f["name"] for f in a["schema"]["fields"]] == ["Lieferant", "Betrag"]

    assert client.get("/assistants").json()[0]["id"] == a["id"]
    assert client.get(f"/assistants/{a['id']}").json()["name"] == "Invoices"

    resp = client.put(
        f"/assistants/{a['id']}",
        json={"name": "Rechnungen", "schema": {"fields": [{"name": "Nur"}]}},
    )
    assert resp.json()["name"] == "Rechnungen"
    assert len(resp.json()["schema"]["fields"]) == 1

    assert client.delete(f"/assistants/{a['id']}").status_code == 204
    assert client.get(f"/assistants/{a['id']}").status_code == 404


def test_create_with_unknown_engine_rejected(client):
    resp = client.post(
        "/assistants", json={"name": "x", "engine": "nope", "schema": SCHEMA}
    )
    assert resp.status_code == 422


def test_create_with_invalid_schema_rejected(client):
    resp = client.post(
        "/assistants", json={"name": "x", "engine": "claude", "schema": {"fields": []}}
    )
    assert resp.status_code == 422


RESULT = ExtractionResult(
    values=[
        FieldValue(field="Lieferant", value="Meridian", confidence="high", needs_review=False),
        FieldValue(field="Betrag", value=8450.0, currency="EUR", confidence="low",
                   needs_review=True),
    ],
    engine="claude", model="m", usage={},
)


def _fake_bulk(documents, schema, engine="claude", model=None, on_update=None, **kw):
    entries = [
        BulkFileEntry(
            filename=name, status="done", result=RESULT, needs_review=True, duration_s=0.5,
            cost=CostBreakdown(engine=engine, model="m", input_tokens=10, output_tokens=5,
                               pricing_known=True, total_cost=0.02),
        )
        for name, _ in documents
    ]
    report = BulkReport(status="done", engine=engine, model="m", total=len(entries),
                        completed=len(entries), needs_review_files=len(entries),
                        total_cost_usd=0.02 * len(entries), entries=entries)
    if on_update:
        on_update(report)
    return report


@pytest.fixture
def fake_bulk(monkeypatch):
    monkeypatch.setattr(docstill, "bulk_extract", _fake_bulk)
    monkeypatch.setattr(docstill, "get_engine", lambda *a, **k: object())


def _wait_results(client, assistant_id, n, timeout=3.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        rows = client.get(f"/assistants/{assistant_id}/results").json()
        if len(rows) >= n:
            return rows
        time.sleep(0.02)
    raise AssertionError("results were not persisted in time")


def test_assistant_bulk_persists_results(client, fake_bulk):
    a = _create(client)
    resp = client.post(
        f"/assistants/{a['id']}/bulk",
        files=[
            ("files", ("a.pdf", PDF, "application/pdf")),
            ("files", ("b.pdf", PDF, "application/pdf")),
        ],
    )
    assert resp.status_code == 202
    job_id = resp.json()["job_id"]
    assert client.get(f"/bulk/{job_id}").status_code == 200

    rows = _wait_results(client, a["id"], 2)
    assert {r["filename"] for r in rows} == {"a.pdf", "b.pdf"}
    assert rows[0]["values"][1]["needs_review"] is True

    listed = client.get(f"/assistants/{a['id']}").json()
    assert listed["doc_count"] == 2
    assert listed["review_count"] == 2
    assert listed["total_cost_usd"] == pytest.approx(0.04)

    review = client.get(f"/assistants/{a['id']}/results", params={"filter": "review"}).json()
    assert len(review) == 2


def test_bulk_unknown_assistant_404(client, fake_bulk):
    resp = client.post(
        "/bulk-nonexistent" if False else "/assistants/nope/bulk",
        files=[("files", ("a.pdf", PDF, "application/pdf"))],
    )
    assert resp.status_code == 404


def test_export_xlsx(client, fake_bulk):
    a = _create(client)
    client.post(
        f"/assistants/{a['id']}/bulk",
        files=[("files", ("a.pdf", PDF, "application/pdf"))],
    )
    _wait_results(client, a["id"], 1)

    resp = client.get(f"/assistants/{a['id']}/export.xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["Filename", "Status", "Needs review", "Lieferant", "Betrag",
                      "Cost USD", "Extracted at"]
    row = [c.value for c in ws[2]]
    assert row[0] == "a.pdf"
    assert row[3] == "Meridian"
    assert row[4] == 8450.0
    # low-confidence Betrag cell carries the review fill
    assert ws.cell(row=2, column=5).fill.start_color.rgb.endswith("FCF6E7")


def test_export_unknown_assistant_404(client):
    assert client.get("/assistants/nope/export.xlsx").status_code == 404
